# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.18.1
#   kernelspec:
#     display_name: Planning_Agent_NDD
#     language: python
#     name: python3
# ---

# %% [markdown]
# # The Audit Planning Agent - Draft Notebook

# %% [markdown]
# ## Stage 1 - Ingestion
# Ingesting trial balance before mapping into a full fledge financial report


# %%
import pandas as pd
import numpy as np
from IPython.display import display
from pathlib import Path
import yaml
from typing import Dict, List, Optional, Any

# %%
# Determine notebook directory for path resolution
try:
    NOTEBOOK_DIR = Path(__file__).resolve().parent
    REPO_ROOT = NOTEBOOK_DIR.parent
except NameError:  # pragma: no cover - jupyter magic
    # When running in Jupyter, find repo root by looking for data/configs dirs
    current = Path.cwd()
    # Check if we're in notebooks directory - if so, go up one level
    if current.name == "notebooks":
        REPO_ROOT = current.parent
    # Check if data and configs exist in current directory
    elif (current / "data").exists() and (current / "configs").exists():
        REPO_ROOT = current
    # Otherwise, try parent directory
    elif (current.parent / "data").exists() and (current.parent / "configs").exists():
        REPO_ROOT = current.parent
    else:
        # Fallback: assume current is repo root
        REPO_ROOT = current
    NOTEBOOK_DIR = REPO_ROOT / "notebooks"

DATA_DIR = REPO_ROOT / "data"
CONFIGS_DIR = REPO_ROOT / "configs"

# %%
# Stage 1 - TB Ingestion
# Load trial balance file
tb_path = DATA_DIR / "TB_2025.xlsx"
tb_raw = pd.read_excel(tb_path)

print(f"Loaded TB: {tb_raw.shape[0]} rows, {tb_raw.shape[1]} columns")
print(f"Columns: {list(tb_raw.columns)}")
display(tb_raw.head())

# %%
# Enrich TB with helper fields and compute measurement columns
tb = tb_raw.copy()

# Normalize account_id to string for consistent matching
tb["account_code_str"] = tb["account_id"].astype(str)

# Add category flags
tb["is_bs"] = tb["Category"] == "BS"
tb["is_pl"] = tb["Category"] == "PL"

# Compute measurement columns:
# - BS accounts: use closing balance (Closing Dr - Closing Cr)
# - PL accounts: use period movement (Dr - Cr)
tb["closing_balance"] = np.where(
    tb["is_bs"],
    tb["Closing Dr"].fillna(0) - tb["Closing Cr"].fillna(0),
    np.nan,
)
tb["period_movement"] = np.where(
    tb["is_pl"],
    tb["Dr"].fillna(0) - tb["Cr"].fillna(0),
    np.nan,
)

# Create unified measurement column
tb["amount"] = tb["closing_balance"].fillna(0) + tb["period_movement"].fillna(0)

print(f"\nTB enriched with measurement columns")
print(f"BS accounts: {tb['is_bs'].sum()}, PL accounts: {tb['is_pl'].sum()}")
display(tb[["account_code_str", "Description", "Category", "amount"]].head(10))

# %%
# Stage 2 - Mapping Leaf Accounts
# Load VAS mapping configuration
mapping_path = CONFIGS_DIR / "vas_account_mapping.yaml"
with open(mapping_path, "r", encoding="utf-8") as f:
    mapping_config = yaml.safe_load(f)

mapping_rules = mapping_config["mapping_rules"]
account_mappings = mapping_rules["account_mappings"]
match_order = mapping_rules["hierarchy"]["match_order"]

print(f"Loaded mapping config: {mapping_config['version']}")
print(f"Match order: {match_order}")

# %%
# Build lookup dictionaries for each hierarchy level
sub_level_map: Dict[str, List[Dict[str, Any]]] = account_mappings.get("sub_level", {})
second_level_map: Dict[str, List[Dict[str, Any]]] = account_mappings.get(
    "second_level", {}
)
first_level_map: Dict[str, List[Dict[str, Any]]] = account_mappings.get(
    "first_level", {}
)

print(f"Mapping dictionaries loaded:")
print(f"  - Sub-level: {len(sub_level_map)} entries")
print(f"  - Second-level: {len(second_level_map)} entries")
print(f"  - First-level: {len(first_level_map)} entries")

# %%
# Identify leaf accounts (accounts with no children in the TB)
# A leaf account is one where no other account_code_str is longer and starts with it
tb["is_leaf"] = False

for idx, row in tb.iterrows():
    account_code = row["account_code_str"]
    # Check if any other account is longer and starts with this account code
    longer_accounts = tb["account_code_str"].str.len() > len(account_code)
    starts_with = tb["account_code_str"].str.startswith(account_code)
    has_children = (longer_accounts & starts_with).any()
    tb.loc[idx, "is_leaf"] = not has_children

leaf_accounts = tb[tb["is_leaf"]].copy()
print(f"\nIdentified {len(leaf_accounts)} leaf accounts out of {len(tb)} total accounts")
display(leaf_accounts[["account_code_str", "Description", "Category", "amount"]].head(10))

# %%
# Apply longest-match mapping to leaf accounts
def find_mapping(account_code: str) -> Optional[Dict[str, Any]]:
    """
    Find mapping for account code using longest-match strategy.
    
    Strategy:
    1. Try exact match at each hierarchy level (sub → second → first)
    2. If no exact match, try prefix matching: find the longest key that is a
       prefix of the account_code (e.g., '13111' matches '1311' in sub_level_map)
    3. This handles cases where TB has more granular accounts than the mapping config
    """
    # Try sub-level: exact match first
    if account_code in sub_level_map:
        mappings = sub_level_map[account_code]
        if mappings:
            return mappings[0]  # Take first mapping if multiple
    
    # Try sub-level: prefix match (longest prefix wins)
    # Match when mapping key is a prefix of account_code (e.g., '1311' matches '13111')
    prefix_matches = [
        k for k in sub_level_map.keys()
        if account_code.startswith(str(k))
    ]
    if prefix_matches:
        # Use longest matching prefix to maintain longest-match strategy
        longest_prefix = max(prefix_matches, key=lambda x: len(str(x)))
        mappings = sub_level_map[longest_prefix]
        if mappings:
            return mappings[0]

    # Try second-level: exact match first
    if len(account_code) >= 4:
        second_code = account_code[:4]
        if second_code in second_level_map:
            mappings = second_level_map[second_code]
            if mappings:
                return mappings[0]
        
        # Try second-level: prefix match
        prefix_matches = [
            k for k in second_level_map.keys()
            if account_code.startswith(str(k))
        ]
        if prefix_matches:
            longest_prefix = max(prefix_matches, key=lambda x: len(str(x)))
            mappings = second_level_map[longest_prefix]
            if mappings:
                return mappings[0]

    # Try first-level: exact match first
    if len(account_code) >= 3:
        first_code = account_code[:3]
        if first_code in first_level_map:
            mappings = first_level_map[first_code]
            if mappings:
                return mappings[0]
        
        # Try first-level: prefix match
        prefix_matches = [
            k for k in first_level_map.keys()
            if account_code.startswith(str(k))
        ]
        if prefix_matches:
            longest_prefix = max(prefix_matches, key=lambda x: len(str(x)))
            mappings = first_level_map[longest_prefix]
            if mappings:
                return mappings[0]

    return None


# Apply mapping to each leaf account
mapped_results = []
unmapped_accounts = []

for idx, row in leaf_accounts.iterrows():
    account_code = row["account_code_str"]
    mapping = find_mapping(account_code)

    result = {
        "account_code": account_code,
        "description": row["Description"],
        "category": row["Category"],
        "amount": row["amount"],
        "mapped": mapping is not None,
    }

    if mapping:
        result["leadsheet_code"] = mapping.get("leadsheet", {}).get("code")
        result["leadsheet_name"] = mapping.get("leadsheet", {}).get("name")
        result["fs_code"] = mapping.get("fs_target", {}).get("code")
        result["fs_label"] = mapping.get("fs_target", {}).get("label")
    else:
        result["leadsheet_code"] = None
        result["leadsheet_name"] = None
        result["fs_code"] = None
        result["fs_label"] = None
        unmapped_accounts.append(account_code)

    mapped_results.append(result)

mapped_df = pd.DataFrame(mapped_results)
print(f"\nMapping complete:")
print(f"  - Mapped: {mapped_df['mapped'].sum()} accounts")
print(f"  - Unmapped: {len(unmapped_accounts)} accounts")

if unmapped_accounts:
    print(f"\nUnmapped accounts: {unmapped_accounts[:10]}")

display(mapped_df.head(100))

# %%
# Stage 3 - Aggregated Outputs
# Aggregate by FS target (for financial statement lines)
fs_summary = (
    mapped_df[mapped_df["mapped"]]
    .groupby(["fs_code", "fs_label", "category"], as_index=False)
    .agg({"amount": "sum"})
    .sort_values(["category", "fs_code"])
)

print(f"\nFS Summary by target line:")
print(f"  - {len(fs_summary)} unique FS lines")
display(fs_summary.head(20))

# %%
# Aggregate by leadsheet (for working papers)
leadsheet_summary = (
    mapped_df[mapped_df["mapped"] & mapped_df["leadsheet_code"].notna()]
    .groupby(["leadsheet_code", "leadsheet_name", "category"], as_index=False)
    .agg({"amount": "sum"})
    .sort_values(["category", "leadsheet_code"])
)

print(f"\nLeadsheet Summary:")
print(f"  - {len(leadsheet_summary)} unique leadsheets")
display(leadsheet_summary.head(20))

# %%
# Display balance sheet vs P&L breakdown
bs_total = mapped_df[mapped_df["category"] == "BS"]["amount"].sum()
pl_total = mapped_df[mapped_df["category"] == "PL"]["amount"].sum()

print(f"\n=== Summary Totals ===")
print(f"Balance Sheet total: {bs_total:,.0f}")
print(f"Profit & Loss total: {pl_total:,.0f}")
print(f"Net difference: {bs_total + pl_total:,.0f}")

# %%
# Stage 4 - Populate Financial Statement Templates
TEMPLATES_DIR = REPO_ROOT / "templates"
ARTIFACTS_DIR = REPO_ROOT / "artifacts"

# %%
# Load template structures
def load_template(template_path: Path) -> pd.DataFrame:
    """Load template Excel file and return as DataFrame."""
    return pd.read_excel(template_path, header=None)


def extract_template_structure(template_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract template structure: row index, label (col B), FS code (col C).
    Returns DataFrame with columns: row_idx, label, fs_code.
    """
    structure = []
    for idx, row in template_df.iterrows():
        label = row[1] if pd.notna(row[1]) else None
        fs_code = row[2] if pd.notna(row[2]) else None
        
        # Skip header rows and empty rows
        if fs_code is not None and str(fs_code).strip() not in ["", "Code", "nan"]:
            structure.append({
                "row_idx": idx,
                "label": str(label).strip() if label is not None else "",
                "fs_code": str(fs_code).strip(),
            })
    
    return pd.DataFrame(structure)


def build_parent_child_hierarchy(template_structure: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Build parent-child hierarchy from template structure.
    Returns dict mapping parent_code -> [child_codes].
    
    Hierarchy rules:
    - BS codes: 100-level -> 110-level -> 111-level (numeric grouping)
    - IS codes: Mostly calculated, but some parent-child relationships exist
    - Special codes with suffixes (411a, 421a) are children of base codes (411, 421)
    """
    parent_child = {}
    codes = template_structure["fs_code"].tolist()
    
    # Helper to check if code is numeric
    def is_numeric(code: str) -> bool:
        try:
            float(code)
            return True
        except (ValueError, TypeError):
            return False
    
    # Helper to get numeric prefix (e.g., "111" from "111", "110" from "110")
    def get_numeric_prefix(code: str, prefix_len: int) -> Optional[str]:
        if not is_numeric(code) or len(code) < prefix_len:
            return None
        return code[:prefix_len]
    
    for code in codes:
        children = []
        
        # Handle special suffix codes (411a, 421a, etc.)
        if not code[-1].isdigit() and len(code) > 3:
            # This is a suffix code like "411a" or "421a"
            base_code = code[:-1]  # Remove suffix
            if base_code in codes:
                # The base code is the parent, but we're looking for children of base_code
                # So we'll handle this in the base code iteration
                continue
        
        # For numeric codes, find children based on hierarchy levels
        if is_numeric(code):
            code_num = float(code)
            
            # BS hierarchy patterns:
            # - 100-level (100, 200, 300, 400) -> 110-level children (110-199, 210-299, etc.)
            # - 110-level (110, 120, 130, etc.) -> 111-level children (111, 112, 121, etc.)
            # - 111-level are typically leaf nodes
            
            # Check for 100-level parent (e.g., 100, 200, 300, 400)
            # Exclude calculated totals: 270, 440
            if code_num >= 100 and code_num < 1000 and code_num % 100 == 0:
                # Find children in the next hundred range (e.g., 100 -> 110-199)
                range_start = int(code_num)
                range_end = range_start + 100
                excluded_totals = [270, 440]  # These are calculated, not children
                for c in codes:
                    if c != code and is_numeric(c):
                        c_num = float(c)
                        # Exclude calculated totals and ensure it's a direct child (ends in 0, not 00)
                        if (range_start < c_num < range_end 
                            and c_num % 10 == 0 
                            and c_num % 100 != 0
                            and c_num not in excluded_totals):
                            children.append(c)
            
            # Check for 110-level parent (e.g., 110, 120, 130) -> 111-level children
            elif code_num >= 110 and code_num < 1000 and code_num % 10 == 0 and code_num % 100 != 0:
                # Find children that start with this code (e.g., 110 -> 111, 112)
                for c in codes:
                    if c != code and is_numeric(c) and c.startswith(code):
                        children.append(c)
            
            # IS hierarchy: simpler, mostly 2-digit codes
            # 01, 02 are base, 10 is calculated from them
            # But we'll handle IS calculated lines separately
            
        # Handle suffix codes as children of base codes
        # (e.g., 411a, 411b are children of 411)
        base_code = code[:-1] if len(code) > 3 and not code[-1].isdigit() else None
        if base_code and base_code in codes:
            if base_code not in parent_child:
                parent_child[base_code] = []
            if code not in parent_child[base_code]:
                parent_child[base_code].append(code)
        
        if children:
            parent_child[code] = sorted(children)
    
    return parent_child


# Load Balance Sheet template
bs_template = load_template(TEMPLATES_DIR / "BS_Template.xlsx")
bs_structure = extract_template_structure(bs_template)
bs_hierarchy = build_parent_child_hierarchy(bs_structure)

print(f"\n=== Balance Sheet Template Structure ===")
print(f"Total rows in template: {len(bs_template)}")
print(f"Rows with FS codes: {len(bs_structure)}")
print(f"Parent-child relationships: {len(bs_hierarchy)}")
print(f"\nSample hierarchy (first 5):")
for parent, children in list(bs_hierarchy.items())[:5]:
    print(f"  {parent} -> {children}")

# Load Income Statement template
is_template = load_template(TEMPLATES_DIR / "Income_Statement_Template.xlsx")
is_structure = extract_template_structure(is_template)
is_hierarchy = build_parent_child_hierarchy(is_structure)

print(f"\n=== Income Statement Template Structure ===")
print(f"Total rows in template: {len(is_template)}")
print(f"Rows with FS codes: {len(is_structure)}")
print(f"Parent-child relationships: {len(is_hierarchy)}")
print(f"\nSample hierarchy (first 5):")
for parent, children in list(is_hierarchy.items())[:5]:
    print(f"  {parent} -> {children}")

# %%
# Map amounts from mapped_df to template structures
def populate_leaf_amounts(
    template_df: pd.DataFrame,
    template_structure: pd.DataFrame,
    mapped_data: pd.DataFrame,
    amount_col: str = "amount",
) -> pd.DataFrame:
    """
    Populate template with amounts from mapped data.
    Returns template DataFrame with amounts in column 3 (Current year Unaudited).
    """
    result_df = template_df.copy()
    
    # Create lookup: fs_code -> amount (sum if multiple accounts map to same code)
    mapped_bs = mapped_data[mapped_data["category"] == "BS"].copy()
    mapped_pl = mapped_data[mapped_data["category"] == "PL"].copy()
    
    # Aggregate by fs_code for BS and PL separately
    bs_amounts = (
        mapped_bs[mapped_bs["mapped"] & mapped_bs["fs_code"].notna()]
        .groupby("fs_code", as_index=False)
        .agg({amount_col: "sum"})
    )
    bs_amount_dict = dict(zip(bs_amounts["fs_code"].astype(str), bs_amounts[amount_col]))
    
    pl_amounts = (
        mapped_pl[mapped_pl["mapped"] & mapped_pl["fs_code"].notna()]
        .groupby("fs_code", as_index=False)
        .agg({amount_col: "sum"})
    )
    pl_amount_dict = dict(zip(pl_amounts["fs_code"].astype(str), pl_amounts[amount_col]))
    
    # Determine which template we're working with based on FS codes
    # BS codes are 100+, IS codes are 01-71
    all_amounts = {**bs_amount_dict, **pl_amount_dict}
    
    # Populate amounts in column 3
    for idx, row in template_structure.iterrows():
        fs_code = str(row["fs_code"]).strip()
        template_row_idx = row["row_idx"]
        
        if fs_code in all_amounts:
            result_df.loc[template_row_idx, 3] = all_amounts[fs_code]
    
    return result_df


# Populate leaf amounts
bs_template_populated = populate_leaf_amounts(bs_template, bs_structure, mapped_df)
is_template_populated = populate_leaf_amounts(is_template, is_structure, mapped_df)

print(f"\n=== Populated Leaf Amounts ===")
print(f"Balance Sheet: populated {bs_structure['fs_code'].isin([str(k) for k in mapped_df[mapped_df['category'] == 'BS']['fs_code'].dropna().astype(str).unique()]).sum()} lines")
print(f"Income Statement: populated {is_structure['fs_code'].isin([str(k) for k in mapped_df[mapped_df['category'] == 'PL']['fs_code'].dropna().astype(str).unique()]).sum()} lines")

# %%
# Calculate parent/aggregate lines
def calculate_parent_amounts(
    template_df: pd.DataFrame,
    template_structure: pd.DataFrame,
    hierarchy: Dict[str, List[str]],
) -> pd.DataFrame:
    """
    Calculate parent line amounts by summing child lines.
    Updates template_df in place for parent codes.
    """
    result_df = template_df.copy()
    
    # Process parents in order (process children before parents)
    # Sort by code length descending to process deeper levels first
    parent_codes = sorted(hierarchy.keys(), key=lambda x: (len(x), x), reverse=True)
    
    for parent_code in parent_codes:
        children = hierarchy[parent_code]
        
        # Find parent row
        parent_row = template_structure[template_structure["fs_code"] == parent_code]
        if parent_row.empty:
            continue
        
        parent_idx = parent_row.iloc[0]["row_idx"]
        
        # Sum child amounts
        total = 0.0
        for child_code in children:
            child_row = template_structure[template_structure["fs_code"] == child_code]
            if not child_row.empty:
                child_idx = child_row.iloc[0]["row_idx"]
                child_amount = result_df.loc[child_idx, 3]
                if pd.notna(child_amount):
                    total += float(child_amount)
        
        # Set parent amount
        result_df.loc[parent_idx, 3] = total
    
    return result_df


# Calculate parent amounts for BS and IS
bs_template_with_parents = calculate_parent_amounts(
    bs_template_populated, bs_structure, bs_hierarchy
)
is_template_with_parents = calculate_parent_amounts(
    is_template_populated, is_structure, is_hierarchy
)

print(f"\n=== Calculated Parent Amounts ===")
print(f"Balance Sheet: calculated {len(bs_hierarchy)} parent lines")
print(f"Income Statement: calculated {len(is_hierarchy)} parent lines")

# %%
# Calculate derived lines (Income Statement formulas)
def calculate_is_derived_lines(
    template_df: pd.DataFrame,
    template_structure: pd.DataFrame,
) -> pd.DataFrame:
    """
    Calculate Income Statement derived lines using formulas:
    - 10 = 01 - 02 (Net revenue)
    - 20 = 10 - 11 (Gross profit)
    - 30 = 20 + 21 - 22 - 25 - 26 (Net operating profit)
    - 40 = (calculated if needed)
    - 50 = 30 + 31 - 32 - 40 (Net profit before tax)
    - 60 = 50 - 51 - 52 (Net profit after tax)
    """
    result_df = template_df.copy()
    
    def get_amount(fs_code: str) -> float:
        """Get amount for a given FS code."""
        row = template_structure[template_structure["fs_code"] == fs_code]
        if row.empty:
            return 0.0
        idx = row.iloc[0]["row_idx"]
        amount = result_df.loc[idx, 3]
        return float(amount) if pd.notna(amount) else 0.0
    
    def set_amount(fs_code: str, amount: float):
        """Set amount for a given FS code."""
        row = template_structure[template_structure["fs_code"] == fs_code]
        if row.empty:
            return
        idx = row.iloc[0]["row_idx"]
        result_df.loc[idx, 3] = amount
    
    # 10 = 01 - 02 (Net revenue)
    set_amount("10", get_amount("01") - get_amount("02"))
    
    # 20 = 10 - 11 (Gross profit)
    set_amount("20", get_amount("10") - get_amount("11"))
    
    # 30 = 20 + 21 - 22 - 25 - 26 (Net operating profit)
    set_amount("30", get_amount("20") + get_amount("21") - get_amount("22") - get_amount("25") - get_amount("26"))
    
    # 40 = Other loss (if exists, keep as is or calculate)
    # Usually this is a direct line, not calculated
    
    # 50 = 30 + 31 - 32 - 40 (Net profit before tax)
    set_amount("50", get_amount("30") + get_amount("31") - get_amount("32") - get_amount("40"))
    
    # 60 = 50 - 51 - 52 (Net profit after tax)
    set_amount("60", get_amount("50") - get_amount("51") - get_amount("52"))
    
    # BS totals: 270 = 100 + 200, 440 = 300 + 400
    # Note: These are BS codes, but we'll handle them if IS template doesn't have them
    # Actually, these should be in BS template, not IS
    
    return result_df


# Calculate BS totals
def calculate_bs_totals(
    template_df: pd.DataFrame,
    template_structure: pd.DataFrame,
) -> pd.DataFrame:
    """
    Calculate Balance Sheet totals:
    - 270 = 100 + 200 (Total Assets)
    - 440 = 300 + 400 (Total Resources)
    """
    result_df = template_df.copy()
    
    def get_amount(fs_code: str) -> float:
        """Get amount for a given FS code."""
        row = template_structure[template_structure["fs_code"] == fs_code]
        if row.empty:
            return 0.0
        idx = row.iloc[0]["row_idx"]
        amount = result_df.loc[idx, 3]
        return float(amount) if pd.notna(amount) else 0.0
    
    def set_amount(fs_code: str, amount: float):
        """Set amount for a given FS code."""
        row = template_structure[template_structure["fs_code"] == fs_code]
        if row.empty:
            return
        idx = row.iloc[0]["row_idx"]
        result_df.loc[idx, 3] = amount
    
    # 270 = 100 + 200 (Total Assets)
    set_amount("270", get_amount("100") + get_amount("200"))
    
    # 440 = 300 + 400 (Total Resources)
    set_amount("440", get_amount("300") + get_amount("400"))
    
    return result_df


# Calculate derived lines
is_template_final = calculate_is_derived_lines(is_template_with_parents, is_structure)
bs_template_final = calculate_bs_totals(bs_template_with_parents, bs_structure)

print(f"\n=== Calculated Derived Lines ===")
print("Income Statement: calculated lines 10, 20, 30, 50, 60")
print("Balance Sheet: calculated totals 270, 440")

# %%
# Create summary DataFrames for display
def create_fs_dataframe(
    template_df: pd.DataFrame,
    template_structure: pd.DataFrame,
    statement_name: str,
) -> pd.DataFrame:
    """
    Create a clean DataFrame for display with key columns.
    """
    rows = []
    for _, struct_row in template_structure.iterrows():
        row_idx = struct_row["row_idx"]
        label = struct_row["label"]
        fs_code = struct_row["fs_code"]
        amount = template_df.loc[row_idx, 3] if pd.notna(template_df.loc[row_idx, 3]) else 0.0
        
        rows.append({
            "FS_Code": fs_code,
            "Label": label,
            "Amount": amount,
        })
    
    df = pd.DataFrame(rows)
    df["Statement"] = statement_name
    return df


bs_display_df = create_fs_dataframe(bs_template_final, bs_structure, "Balance Sheet")
is_display_df = create_fs_dataframe(is_template_final, is_structure, "Income Statement")

print(f"\n=== Financial Statement DataFrames ===")
print(f"\nBalance Sheet ({len(bs_display_df)} lines):")
display(bs_display_df.head(30))

print(f"\nIncome Statement ({len(is_display_df)} lines):")
display(is_display_df)

# %%
# Write populated templates to Excel files
ARTIFACTS_DIR.mkdir(exist_ok=True)

bs_output_path = ARTIFACTS_DIR / "fs_balance_sheet_populated.xlsx"
is_output_path = ARTIFACTS_DIR / "fs_income_statement_populated.xlsx"

# Write using openpyxl to preserve formatting (if available) or use pandas
try:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Load original template to preserve formatting
    bs_wb = load_workbook(TEMPLATES_DIR / "BS_Template.xlsx")
    bs_ws = bs_wb.active
    
    # Update amounts in column D (index 4, 1-based)
    for _, struct_row in bs_structure.iterrows():
        row_idx = struct_row["row_idx"] + 1  # Convert to 1-based
        fs_code = struct_row["fs_code"]
        amount = bs_template_final.loc[struct_row["row_idx"], 3]
        if pd.notna(amount):
            bs_ws.cell(row=row_idx, column=4, value=float(amount))
    
    bs_wb.save(bs_output_path)
    print(f"\n✓ Saved Balance Sheet to: {bs_output_path}")
    
    # Same for Income Statement
    is_wb = load_workbook(TEMPLATES_DIR / "Income_Statement_Template.xlsx")
    is_ws = is_wb.active
    
    for _, struct_row in is_structure.iterrows():
        row_idx = struct_row["row_idx"] + 1  # Convert to 1-based
        fs_code = struct_row["fs_code"]
        amount = is_template_final.loc[struct_row["row_idx"], 3]
        if pd.notna(amount):
            is_ws.cell(row=row_idx, column=4, value=float(amount))
    
    is_wb.save(is_output_path)
    print(f"✓ Saved Income Statement to: {is_output_path}")
    
except ImportError:
    # Fallback to pandas if openpyxl not available
    bs_template_final.to_excel(bs_output_path, index=False, header=False)
    is_template_final.to_excel(is_output_path, index=False, header=False)
    print(f"\n✓ Saved Balance Sheet to: {bs_output_path} (pandas fallback)")
    print(f"✓ Saved Income Statement to: {is_output_path} (pandas fallback)")

# %%
# Display reconciliation totals
bs_total_assets = bs_display_df[bs_display_df["FS_Code"] == "270"]["Amount"].values[0] if len(bs_display_df[bs_display_df["FS_Code"] == "270"]) > 0 else 0
bs_total_resources = bs_display_df[bs_display_df["FS_Code"] == "440"]["Amount"].values[0] if len(bs_display_df[bs_display_df["FS_Code"] == "440"]) > 0 else 0
is_net_profit = is_display_df[is_display_df["FS_Code"] == "60"]["Amount"].values[0] if len(is_display_df[is_display_df["FS_Code"] == "60"]) > 0 else 0

print(f"\n=== Reconciliation Totals ===")
print(f"Balance Sheet - Total Assets (270): {bs_total_assets:,.0f}")
print(f"Balance Sheet - Total Resources (440): {bs_total_resources:,.0f}")
print(f"Balance Sheet - Difference: {bs_total_assets - bs_total_resources:,.0f}")
print(f"\nIncome Statement - Net Profit After Tax (60): {is_net_profit:,.0f}")
